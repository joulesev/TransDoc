import streamlit as st
import google.generativeai as genai
import pandas as pd
import io
import openpyxl

# --- CONFIGURACIÓN DE PANDAS ---
pd.set_option('display.max_rows', None)
pd.set_option('display.max_colwidth', None)

# --- CONFIGURACIÓN DE LA PÁGINA ---
st.set_page_config(
    page_title="Taller de Datos RAG",
    page_icon="🛠️",
    layout="wide"
)

# --- CONFIGURACIÓN DE LA API DE GEMINI ---
try:
    genai.configure(api_key=st.secrets["GEMINI_API_KEY"])
except (KeyError, FileNotFoundError):
    st.error("🚨 ¡Error de configuración! No se encontró la clave de API de Gemini.")
    st.warning("Por favor, configura el 'Secreto' de Streamlit llamado `GEMINI_API_KEY`.")
    st.stop()

# --- FUNCIÓN DE PREPROCESAMIENTO DE EXCEL (CORREGIDA) ---
def preprocess_excel_merges(uploaded_file):
    """
    Lee un archivo Excel en memoria, detecta las celdas combinadas,
    rellena los valores y devuelve el archivo corregido en memoria.
    """
    # Carga el libro de trabajo, forzando el modo que permite leer la estructura de celdas combinadas.
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Crea una copia de la lista de rangos para iterar de forma segura
        merged_ranges = list(ws.merged_cells.ranges)
        
        for merged_cell_range in merged_ranges:
            # Des-combina las celdas para poder escribir en ellas
            ws.unmerge_cells(str(merged_cell_range))
            
            # Obtiene la celda superior izquierda que contiene el valor
            top_left_cell = merged_cell_range.min_row, merged_cell_range.min_col
            value = ws.cell(row=top_left_cell[0], column=top_left_cell[1]).value
            
            # Rellena todas las celdas del rango original con ese valor
            for row in range(merged_cell_range.min_row, merged_cell_range.max_row + 1):
                for col in range(merged_cell_range.min_col, merged_cell_range.max_col + 1):
                    ws.cell(row=row, column=col).value = value

    # Guarda el libro de trabajo corregido en un objeto de bytes en memoria
    corrected_file_in_memory = io.BytesIO()
    wb.save(corrected_file_in_memory)
    corrected_file_in_memory.seek(0) # Rebobina al principio del archivo
    return corrected_file_in_memory

# --- INICIALIZACIÓN DEL ESTADO DE LA SESIÓN ---
if 'original_content' not in st.session_state:
    st.session_state.original_content = ""
if 'structured_text' not in st.session_state:
    st.session_state.structured_text = "El resultado estructurado por la IA aparecerá aquí..."
if 'chat_history' not in st.session_state:
    st.session_state.chat_history = []

# --- INTERFAZ DE LA APLICACIÓN ---
st.title("🛠️ Taller de Estructuración de Datos para IA (RAG)")
st.markdown("Sube o pega tu contenido, visualízalo, edítalo y usa un asistente de IA para refinarlo. Finalmente, descarga tu documento en formato Markdown.")

col1, col2, col3 = st.columns([1, 1, 1])

with col1:
    with st.container(border=True):
        st.subheader("1. Carga y Visualiza")
        input_method_tab, file_upload_tab = st.tabs(["Pegar Texto", "Subir Archivo Excel"])

        with input_method_tab:
            text_input = st.text_area("Pega texto sin formato aquí", height=150)
            if text_input:
                st.session_state.original_content = text_input

        with file_upload_tab:
            uploaded_file = st.file_uploader("Sube un archivo .xlsx", type=['xlsx'])
            if uploaded_file:
                try:
                    # Preprocesa el archivo para manejar celdas combinadas
                    corrected_file = preprocess_excel_merges(uploaded_file)
                    
                    xls = pd.ExcelFile(corrected_file)
                    sheet_to_display = st.selectbox("Selecciona una hoja para visualizar", xls.sheet_names)
                    if sheet_to_display:
                        df = pd.read_excel(xls, sheet_name=sheet_to_display)
                        st.dataframe(df)
                        
                        full_excel_text = []
                        for name in xls.sheet_names:
                            sheet_df = pd.read_excel(xls, sheet_name=name)
                            if not sheet_df.empty:
                                full_excel_text.append(f"## Hoja: {name}\n\n{sheet_df.to_markdown(index=False)}\n\n")
                        st.session_state.original_content = "".join(full_excel_text)
                except Exception as e:
                    st.error(f"Error al leer el archivo: {e}")

        if st.button("Procesar y Estructurar", use_container_width=True, type="primary"):
            if st.session_state.original_content:
                with st.spinner("🤖 Estructurando el documento inicial..."):
                    try:
                        model = genai.GenerativeModel('gemini-1.5-flash-latest')
                        prompt = f"""
                        Analiza el siguiente texto y reestructúralo en formato Markdown. 
                        Crea un título, un resumen y secciones lógicas. Resalta los datos clave en negrita.

                        --- TEXTO ORIGINAL ---
                        {st.session_state.original_content}
                        --- FIN DEL TEXTO ---
                        """
                        response = model.generate_content(prompt)
                        st.session_state.structured_text = response.text
                        st.success("¡Documento estructurado!")
                    except Exception as e:
                        st.error(f"Error en la estructuración inicial: {e}")
            else:
                st.warning("Por favor, pega texto o sube un archivo.")

with col2:
    with st.container(border=True):
        st.subheader("2. Edita el Resultado")
        edited_text = st.text_area(
            "Puedes editar el texto directamente aquí",
            value=st.session_state.structured_text,
            height=500,
            key="editor"
        )
        st.session_state.structured_text = edited_text
        st.download_button(
            label="📥 Descargar Archivo .md",
            data=st.session_state.structured_text,
            file_name="documento_estructurado.md",
            mime="text/markdown",
            use_container_width=True
        )

with col3:
    with st.container(border=True):
        st.subheader("3. Asistente de Edición")
        for message in st.session_state.chat_history:
            with st.chat_message(message["role"]):
                st.markdown(message["content"])
        if instruction := st.chat_input("Da una instrucción para editar..."):
            st.session_state.chat_history.append({"role": "user", "content": instruction})
            with st.chat_message("user"):
                st.markdown(instruction)
            with st.spinner("✍️ La IA está editando el documento..."):
                try:
                    model = genai.GenerativeModel('gemini-1.5-flash-latest')
                    prompt = f"""
                    Actúa como un editor de documentos. Tu tarea es modificar el 'DOCUMENTO ACTUAL' basándote en la 'INSTRUCCIÓN DEL USUARIO'.
                    Debes devolver el **documento completo y modificado**, no solo una respuesta a la instrucción.

                    --- DOCUMENTO ACTUAL ---
                    {st.session_state.structured_text}
                    --- FIN DEL DOCUMENTO ---

                    --- INSTRUCCIÓN DEL USUARIO ---
                    {instruction}
                    --- FIN DE LA INSTRUCCIÓN ---
                    """
                    response = model.generate_content(prompt)
                    st.session_state.structured_text = response.text
                    st.session_state.chat_history = []
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al editar con la IA: {e}")
                    
